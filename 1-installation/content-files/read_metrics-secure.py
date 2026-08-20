#!/usr/bin/env python3

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


# --------------------------------------------------
# Configuration
# --------------------------------------------------

CSV_FILE = Path("/opt/zabbix/data/zabbix_metrics_demo.csv")
XLSX_FILE = Path("/opt/zabbix/data/zabbix_metrics_demo.xlsx")

ALLOWED_METRICS = {
    "cpu_usage",
    "memory_usage",
    "disk_usage",
    "active_connections",
    "service_status",
}

HOST_PATTERN = re.compile(r"^[A-Za-z0-9._-]+$")


# --------------------------------------------------
# Helper functions
# --------------------------------------------------

def fail(message):
    """
    Print errors to stderr and return a non-zero exit code.

    This prevents error messages from being interpreted as a normal
    monitoring value.
    """
    print(f"ERROR: {message}", file=sys.stderr)
    sys.exit(1)


def validate_host(host):
    """
    Only allow safe characters in host names.

    Examples allowed:
        web-01
        db-01
        server.example.com
        server_01
    """

    if not host:
        fail("Host cannot be empty")

    if len(host) > 100:
        fail("Host name is too long")

    if not HOST_PATTERN.fullmatch(host):
        fail("Invalid host name")


def validate_metric(metric):
    """
    Only metrics explicitly listed in ALLOWED_METRICS are accepted.
    """

    if metric not in ALLOWED_METRICS:
        fail("Invalid metric")


# --------------------------------------------------
# CSV
# --------------------------------------------------

def read_csv_metric(host, metric):

    if not CSV_FILE.is_file():
        fail("CSV file not found")

    try:
        with CSV_FILE.open(
            mode="r",
            encoding="utf-8",
            newline=""
        ) as file:

            reader = csv.DictReader(file)

            if reader.fieldnames is None:
                fail("CSV file has no header")

            required_columns = {"host"} | ALLOWED_METRICS

            missing_columns = required_columns - set(reader.fieldnames)

            if missing_columns:
                fail("CSV file has missing columns")

            for row in reader:

                if row["host"] == host:

                    value = row.get(metric)

                    if value is None or value == "":
                        fail("Metric value is empty")

                    return value

    except csv.Error:
        fail("Invalid CSV file")

    except OSError:
        fail("Unable to read CSV file")

    fail("Host not found")


# --------------------------------------------------
# Excel
# --------------------------------------------------

def read_xlsx_metric(host, metric):

    if not XLSX_FILE.is_file():
        fail("Excel file not found")

    try:
        workbook = load_workbook(
            XLSX_FILE,
            read_only=True,
            data_only=True
        )

        worksheet = workbook.active

        rows = worksheet.iter_rows(values_only=True)

        try:
            headers = next(rows)
        except StopIteration:
            workbook.close()
            fail("Excel file is empty")

        headers = [
            str(header).strip()
            if header is not None
            else ""
            for header in headers
        ]

        required_columns = {"host"} | ALLOWED_METRICS

        missing_columns = required_columns - set(headers)

        if missing_columns:
            workbook.close()
            fail("Excel file has missing columns")

        host_index = headers.index("host")
        metric_index = headers.index(metric)

        for row in rows:

            if host_index >= len(row):
                continue

            current_host = row[host_index]

            if current_host == host:

                if metric_index >= len(row):
                    workbook.close()
                    fail("Metric value not found")

                value = row[metric_index]

                workbook.close()

                if value is None:
                    fail("Metric value is empty")

                return value

        workbook.close()

    except OSError:
        fail("Unable to read Excel file")

    except Exception:
        fail("Invalid Excel file")

    fail("Host not found")


# --------------------------------------------------
# Main
# --------------------------------------------------

def main():

    if len(sys.argv) != 4:
        fail(
            "Usage: read_metrics.py <csv|xlsx> <host> <metric>"
        )

    source_type = sys.argv[1]
    host = sys.argv[2]
    metric = sys.argv[3]

    # Validate all user-controlled arguments
    validate_host(host)
    validate_metric(metric)

    if source_type == "csv":
        value = read_csv_metric(host, metric)

    elif source_type == "xlsx":
        value = read_xlsx_metric(host, metric)

    else:
        fail("Invalid source type")

    # Important for Zabbix:
    # successful stdout should contain only the monitoring value.
    print(value)


if __name__ == "__main__":
    main()