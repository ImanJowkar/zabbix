#!/usr/bin/env python3
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook

ALLOWED_METRICS = {
    'cpu_usage',
    'memory_usage',
    'disk_usage',
    'active_connections',
    'service_status',
}


if metric not in allowed_metrics:
    print("ZBX_NOTSUPPORTED: Invalid metric")
    sys.exit(1)

def fail(message, code=1):
    print(message, file=sys.stderr)
    raise SystemExit(code)


def read_csv(path, host, metric):
    with path.open('r', encoding='utf-8', newline='') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get('host') == host:
                return row.get(metric)
    return None


def read_xlsx(path, host, metric):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['Metrics']
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    index = {name: i for i, name in enumerate(headers)}
    if 'host' not in index or metric not in index:
        return None
    for row in rows:
        if row[index['host']] == host:
            return row[index[metric]]
    return None


def main():
    if len(sys.argv) != 5:
        fail('Usage: read_metrics.py <csv|xlsx> <file> <host> <metric>')

    file_type, file_name, host, metric = sys.argv[1:]
    path = Path(file_name)

    if metric not in ALLOWED_METRICS:
        fail('Unsupported metric')
    if not path.is_file():
        fail('Data file not found')

    if file_type == 'csv':
        value = read_csv(path, host, metric)
    elif file_type == 'xlsx':
        value = read_xlsx(path, host, metric)
    else:
        fail('Unsupported file type')

    if value is None or value == '':
        fail('Host or metric not found')

    print(value)


if __name__ == '__main__':
    main()
